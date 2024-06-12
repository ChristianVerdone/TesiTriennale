import firebase_admin
import pandas as pd
from firebase_admin import credentials
from google.cloud import storage
from openpyxl import load_workbook

import utilsP
from utilsC import splitCel, createItemConto, splitCelSimple


def processNew():
    cred = credentials.Certificate('C:/UTILS/tesitriennale-4d2f1-firebase-adminsdk-2u5v6-69e53a1fdf.json')
    app = firebase_admin.initialize_app(cred, {'storageBucket': 'gs://tesitriennale-4d2f1.appspot.com'}, name='app2')
    bucket_name = 'tesitriennale-4d2f1.appspot.com'
    storage_client = storage.Client.from_service_account_json(
        'C:/UTILS/tesitriennale-4d2f1-firebase-adminsdk-2u5v6-69e53a1fdf.json')
    bucket = storage_client.get_bucket(bucket_name)
    blob = bucket.get_blob('source.xlsx')
    file_bytes = blob.download_as_bytes()
    print("processnew"+blob.name)

    # Save the downloaded bytes to a file
    with open('source.xlsx', 'wb') as f:
        f.write(file_bytes)

    # wb = xlrd.open_workbook(file_contents=file_bytes, encoding_override='ISO-8859-1')
    wb = load_workbook("source.xlsx")
    # sheet = wb.sheet_by_index(1)  # prendo il foglio di lavoro che mi interessa lavorare
    # sheet = wb.active
    # print(sheet['A1'].value)
    listItem = []
    for sheetname in wb.sheetnames:
        if not (sheetname == "Foglio1" or sheetname == "Foglio2"):
            sheet = wb[sheetname]
            print(sheet.title)
            # Ottieni il numero di righe nel secondo foglio
            num_rows = sheet.max_row

            # Stampa il risultato
            # print(f"Numero di righe nel secondo foglio: {num_rows}")
            # wb = xlrd.open_workbook(location)  # copio un workbook identico al file

            # Seleziona il foglio di lavoro con indice 1 (il secondo foglio)
            # sheet = wb.sheet_by_index(1)  # prendo il foglio di lavoro che mi interessa lavorare

            print(sheet.cell(row=1, column=1).value)
            codConto = ''
            desConto = ''
            for n in range(1, num_rows):
                temp = sheet.cell(n, 1)
                if temp.value.__class__ == str:
                    if temp.value.__contains__('Codice'):
                        codConto, desConto = splitCel(sheet.cell(n, 1).value)
                        continue
                    if ((temp.value.__contains__('8.') or temp.value.__contains__('9.'))
                            and not (sheet.cell(n, 1).value.__contains__('Codice'))):
                        codConto, desConto = splitCelSimple(sheet.cell(n, 1).value)
                        continue
                    if sheet.cell(n, 1).value.__contains__('DT'):
                        continue
                    if sheet.cell(n, 1).value.__contains__('//'):
                        continue
                elif temp.value is None:
                    continue

                row = sheet[n]
                item = createItemConto(codConto, desConto, row)
                listItem.append(item)

    utilsP.writeNewFile(listItem)
    # leggi il file xlsx in un dataframe
    df = pd.read_excel('out.xlsx', sheet_name='Sheet')
    # Salva il DataFrame come file csv in una variabile di tipo bytes
    csv_bytes = df.to_csv(index=False).encode('utf-8')
    # Salva il file csv in Firebase Storage
    blob = bucket.blob('file.csv')
    blob.upload_from_string(csv_bytes, content_type='application/octet-stream')
    firebase_admin.delete_app(app)
    return '200 ok'


def processNewFull():
    # Carica il file .xlsx esistente
    file_path = "errored.xlsx"
    workbook = load_workbook(file_path)

    # Seleziona il foglio desiderato (ad esempio, il primo foglio)
    sheet = workbook.active

    # Ottieni il numero di righe nel foglio
    num_rows = sheet.max_row

    # Itera al contrario attraverso le righe, se la riga non è vuota, la elimina
    for i in range(num_rows, 0, -1):
        if sheet[i][0].value is not None:
            sheet.delete_rows(i)

    # Salva le modifiche nel file
    workbook.save(file_path)

    cred = credentials.Certificate('C:/UTILS/tesitriennale-4d2f1-firebase-adminsdk-2u5v6-69e53a1fdf.json')
    app = firebase_admin.initialize_app(cred, {'storageBucket': 'gs://tesitriennale-4d2f1.appspot.com'}, name='app2')
    bucket_name = 'tesitriennale-4d2f1.appspot.com'
    storage_client = storage.Client.from_service_account_json(
        'C:/UTILS/tesitriennale-4d2f1-firebase-adminsdk-2u5v6-69e53a1fdf.json')
    bucket = storage_client.get_bucket(bucket_name)
    blob = bucket.get_blob('source.xlsx')
    file_bytes = blob.download_as_bytes()
    # print("processnew"+blob.name)

    # Save the downloaded bytes to a file
    with open('source.xlsx', 'wb') as f:
        f.write(file_bytes)

    # wb = xlrd.open_workbook(file_contents=file_bytes, encoding_override='ISO-8859-1')
    wb = load_workbook("source.xlsx")
    # sheet = wb.sheet_by_index(1)  # prendo il foglio di lavoro che mi interessa lavorare
    # sheet = wb.active
    # print(sheet['A1'].value)
    listItem = []
    for sheetname in wb.sheetnames:
        if sheetname == "Foglio2":
            sheet = wb[sheetname]
            # print(sheet.title)
            # Ottieni il numero di righe nel secondo foglio
            num_rows = sheet.max_row

            # Stampa il risultato
            # print(f"Numero di righe nel secondo foglio: {num_rows}")
            # wb = xlrd.open_workbook(location)  # copio un workbook identico al file

            # Seleziona il foglio di lavoro con indice 1 (il secondo foglio)
            # sheet = wb.sheet_by_index(1)  # prendo il foglio di lavoro che mi interessa lavorare

            # print(sheet.cell(row=1, column=1).value)
            codConto = ''
            desConto = ''
            for n in range(1, num_rows):
                temp = sheet.cell(n, 1)
                if temp.value.__class__ == str:
                    if temp.value.__contains__('Codice'):
                        codConto, desConto = splitCel(sheet.cell(n, 1).value)
                        continue
                    if ((temp.value.__contains__('8.') or temp.value.__contains__('9.'))
                            and not (sheet.cell(n, 1).value.__contains__('Codice'))):
                        codConto, desConto = splitCelSimple(sheet.cell(n, 1).value)
                        continue
                    if sheet.cell(n, 1).value.__contains__('DT'):
                        continue
                    if sheet.cell(n, 1).value.__contains__('//'):
                        continue
                elif temp.value is None:
                    continue

                row = sheet[n]
                item = createItemConto(codConto, desConto, row)
                listItem.append(item)

    utilsP.writeNewFile(listItem)
    # leggi il file xlsx in un dataframe
    df = pd.read_excel('out.xlsx', sheet_name='Sheet')
    # Salva il DataFrame come file csv in una variabile di tipo bytes
    csv_bytes = df.to_csv(index=False).encode('utf-8')
    # Salva il file csv in Firebase Storage
    blob = bucket.blob('file.csv')
    blob.upload_from_string(csv_bytes, content_type='application/octet-stream')

    workbook = load_workbook(file_path)
    # Seleziona il foglio desiderato (ad esempio, il primo foglio)
    sheet = workbook.active
    if sheet.cell(row=1, column=1).value is not None:
        blob = bucket.blob('errored.xlsx')
        blob.upload_from_filename(file_path)
    else:
        # Ottieni il blob
        blob = bucket.blob('errored.xlsx')
        # Elimina il blob
        blob.delete()

    firebase_admin.delete_app(app)
    return '200 ok'


t = processNewFull()
