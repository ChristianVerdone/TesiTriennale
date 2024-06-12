import firebase_admin
from firebase_admin import credentials
from google.cloud import storage
import pandas as pd
from openpyxl import load_workbook
import utilsP
from utilsC import createItemContoLine


def processFixFile():
    cred = credentials.Certificate('C:/UTILS/tesitriennale-4d2f1-firebase-adminsdk-2u5v6-69e53a1fdf.json')
    app = firebase_admin.initialize_app(cred, {'storageBucket': 'gs://tesitriennale-4d2f1.appspot.com'}, name='app2')
    bucket_name = 'tesitriennale-4d2f1.appspot.com'
    storage_client = storage.Client.from_service_account_json(
        'C:/UTILS/tesitriennale-4d2f1-firebase-adminsdk-2u5v6-69e53a1fdf.json')
    bucket = storage_client.get_bucket(bucket_name)
    blob = bucket.get_blob('fix.xlsx')
    file_bytes = blob.download_as_bytes()

    # Save the downloaded bytes to a file
    with open('fix.xlsx', 'wb') as f:
        f.write(file_bytes)

    # Read the xlsx file into a dataframe
    wb = load_workbook("fix.xlsx")
    listItem = []
    for sheetname in wb.sheetnames:
        if sheetname == "Foglio1":
            sheet = wb[sheetname]
            num_rows = sheet.max_row
            codConto = ''
            desConto = ''
            for n in range(1, num_rows):
                row = sheet[n]
                item = createItemContoLine(row)
                listItem.append(item)

    utilsP.writeNewFile(listItem)
    df = pd.read_excel('out.xlsx', sheet_name='Sheet')
    csv_bytes = df.to_csv(index=False).encode('utf-8')
    blob = bucket.blob('fix.csv')
    blob.upload_from_string(csv_bytes, content_type='application/octet-stream')

    firebase_admin.delete_app(app)
    return '200 ok'
