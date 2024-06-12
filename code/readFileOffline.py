import openpyxl

import utilsP
from utilsC import splitCel, createItemConto, splitCelSimple

location = "CERICT MASTR 2022.xlsx"  # prendo il file
workbook = openpyxl.load_workbook(location)
listItem = []
for sheetname in workbook.sheetnames:
    if not (sheetname == "Foglio1" or sheetname == "Foglio2"):
        sheet = workbook[sheetname]

        print(sheet.title)
        # Ottieni il numero di righe nel secondo foglio
        num_rows = sheet.max_row

        # Stampa il risultato
        # print(f"Numero di righe nel secondo foglio: {num_rows}")
        # wb = xlrd.open_workbook(location)  # copio un workbook identico al file

        # Seleziona il foglio di lavoro con indice 1 (il secondo foglio)
        # sheet = wb.sheet_by_index(1)  # prendo il foglio di lavoro che mi interessa lavorare

        print(sheet.cell(row=1, column=1).value)
        # print(sheet.cell_value(0, 0))
        codConto = ''
        desConto = ''
        for n in range(1, num_rows):
            temp = sheet.cell(n, 1)
            if temp.value.__class__ == str:
                if temp.value.__contains__('Codice'):
                    codConto, desConto = splitCel(sheet.cell(n, 1).value)
                    print("des conto " + desConto)
                    print("cod conto " + codConto)
                    continue
                if ((temp.value.__contains__('8.') or temp.value.__contains__('9.'))
                        and not (sheet.cell(n, 1).value.__contains__('Codice'))):
                    codConto, desConto = splitCelSimple(sheet.cell(n, 1).value)
                    print("des conto " + desConto)
                    print("cod conto " + codConto)
                    continue
                if sheet.cell(n, 1).value.__contains__('DT'):
                    continue
                if sheet.cell(n, 1).value.__contains__('//'):
                    continue
            elif temp.value is None:
                continue

            row = sheet[n]
            item = createItemConto(codConto, desConto, row)
            listItem.append(item)
        '''
        print(desConto)
        print(codConto)
        '''
utilsP.writeNewFile(listItem)
