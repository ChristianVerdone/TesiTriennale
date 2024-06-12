import datetime
from datetime import datetime
import openpyxl
from utilsP import ItemConto
# from firebase_admin import credentials


def splitCel(stringa):
    token = stringa.split()  # Elimino gli spazi dal contenuto della cella "stringa"
    codConto = token[2]  # visto che la cella da dividere ha sempre la stessa costruzione, il codice conto è in 3
    desConto = ''  # per la descrizione del conto mi basta unire tutte le stringe successive al codice conto
    for n in range(3, len(token)):
        desConto = desConto + token[n] + ' '
    return codConto, desConto


def splitCelSimple(stringa):
    token = stringa.split()  # Elimino gli spazi dal contenuto della cella "stringa"
    codConto = token[0]  # visto che la cella da dividere ha sempre la stessa costruzione, il codice contro è in 3
    desConto = ''  # per la descrizione del conto mi basta unire tutte le stringe successive al codice conto
    for n in range(1, len(token)):
        desConto = desConto + token[n] + ' '
    return codConto, desConto


def createItemContoLine(row):
    importo = 0
    codC = row[0].value
    des = row[1].value
    dataOp = row[2].value
    desOp = row[3].value
    numDoc = row[4].value
    dataDoc = row[5].value
    if row[6].value.__class__ == int:
        importo = float(row[6].value)
    elif row[6].value.__class__ == str:
        if row[6].value.__contains__('-'):
            if row[6].value.__contains__('D') or row[6].value.__contains__('A'):
                # Carica il file .xlsx esistente
                file_path = "errored.xlsx"
                workbook = openpyxl.load_workbook(file_path)
                # Seleziona il foglio desiderato (ad esempio, il primo foglio)
                sheet = workbook.active
                # Scrivi la riga nella prima riga vuota
                sheet.append(codC, des, row)
                # Salva le modifiche nel file
                workbook.save(file_path)
            else:
                temp = row[6].value
                if isinstance(temp, float):
                    importo = -temp
                    temp = 0
                else:
                    temp = temp.strip().replace('.', '').replace(',', '.')
                    x = str(temp)
                    try:
                        importo = -float(x)
                    except ValueError:
                        file_path = "errored.xlsx"
                        workbook = openpyxl.load_workbook(file_path)
                        # Seleziona il foglio desiderato (ad esempio, il primo foglio)
                        sheet = workbook.active
                        # Scrivi la riga nella prima riga vuota
                        error = list()
                        error.append(codC)
                        error.append(des)
                        error.append(dataOp)
                        error.append(desOp)
                        error.append(numDoc)
                        error.append(dataDoc)
                        error.append(row[6].value)
                        error.append(row[7].value)
                        sheet.append(error)
                        # Salva le modifiche nel file
                        workbook.save(file_path)
                    temp = 0
        else:
            if row[6].value.__contains__('D') or row[6].value.__contains__('A'):
                # Carica il file .xlsx esistente
                file_path = "errored.xlsx"
                workbook = openpyxl.load_workbook(file_path)
                # Seleziona il foglio desiderato (ad esempio, il primo foglio)
                sheet = workbook.active
                # Scrivi la riga nella prima riga vuota
                error = list()
                error.append(codC)
                error.append(des)
                error.append(dataOp)
                error.append(desOp)
                error.append(numDoc)
                error.append(dataDoc)
                error.append(row[6].value)
                error.append(row[7].value)
                sheet.append(error)
                # Salva le modifiche nel file
                workbook.save(file_path)
            else:
                temp = row[6].value
                if isinstance(temp, float):
                    importo = temp
                    temp = 0
                else:
                    temp = temp.strip().replace('.', '').replace(',', '.')
                    x = str(temp)
                    importo = float(x)
                    temp = 0
    elif row[6].value.__class__ == float:
        importo = row[6].value
    contr = row[7].value
    item = ItemConto(codC, des, dataOp, desOp, numDoc, dataDoc, importo, contr)
    return item


def createItemConto(codC, des, row):
    importo = 0
    # dataOp = datetime.datetime(*xlrd.xldate_as_tuple(row[0], wb.datemode)).strftime("%d/%m/%Y")  # funzione che cattura la data
    if row[0].value.__contains__('//'):
        dataOp = datetime.strptime('01/01/01', "%d/%m/%y")
    else:
        dataOp = datetime.strptime(row[0].value, "%d/%m/%y")
    # cod = row[1]  # codice operazione
    desOp = row[1].value  # descrizione dell'operazione
    numDoc = row[2].value  # numero che indica il documento
    if row[3].value.__contains__('//'):
        dataDoc = datetime.strptime('01/01/01', "%d/%m/%y")
    else:
        dataDoc = datetime.strptime(row[3].value, "%d/%m/%y")  # data di documentazione
    # numFat = row[5] # numero della fattura
    '''
    if not row[6] and not row[7]:  # per la costruzione dell'importo osservo le colonne "dare" e "avere", importo sarà
        importo = 0  # positivo se il float sarà sulla colonna dare altrimenti sarà negativo.
    if row[6]:
        temp = row[6]
        if isinstance(temp, float):
            importo = temp
            temp = 0
        else:
            temp = row[6].strip().replace('.', '').replace(',', '.')
            x = str(temp)
            importo = float(x)
            temp = 0
    elif row[7]:
        temp = row[7]
        if isinstance(temp, float):
            importo = -temp
            temp = 0
        else:
            temp = row[7].strip().replace('.', '').replace(',', '.')
            x = str(temp)
            importo = -float(x)
            temp = 0
    '''
    # print(row[4].value)
    if row[4].value.__class__ == int:
        importo = float(row[4].value)
        # print("importo " + str(importo))
    elif row[4].value.__class__ == str:
        if row[4].value.__contains__('-'):
            if row[4].value.__contains__('D') or row[4].value.__contains__('A'):
                # Carica il file .xlsx esistente
                file_path = "errored.xlsx"
                workbook = openpyxl.load_workbook(file_path)
                # Seleziona il foglio desiderato (ad esempio, il primo foglio)
                sheet = workbook.active
                # Scrivi la riga nella prima riga vuota
                sheet.append(codC, des, row)
                # Salva le modifiche nel file
                workbook.save(file_path)
            else:
                temp = row[4].value
                if isinstance(temp, float):
                    importo = -temp
                    temp = 0
                else:
                    temp = temp.strip().replace('.', '').replace(',', '.')
                    x = str(temp)
                    try:
                        importo = -float(x)
                    except ValueError:
                        file_path = "errored.xlsx"
                        workbook = openpyxl.load_workbook(file_path)
                        # Seleziona il foglio desiderato (ad esempio, il primo foglio)
                        sheet = workbook.active
                        # Scrivi la riga nella prima riga vuota
                        error = list()
                        error.append(codC)
                        error.append(des)
                        error.append(dataOp)
                        error.append(desOp)
                        error.append(numDoc)
                        error.append(dataDoc)
                        error.append(row[4].value)
                        error.append(row[5].value)
                        sheet.append(error)
                        # Salva le modifiche nel file
                        workbook.save(file_path)
                    temp = 0
        else:
            if row[4].value.__contains__('D') or row[4].value.__contains__('A'):
                # Carica il file .xlsx esistente
                file_path = "errored.xlsx"
                workbook = openpyxl.load_workbook(file_path)
                # Seleziona il foglio desiderato (ad esempio, il primo foglio)
                sheet = workbook.active
                # Scrivi la riga nella prima riga vuota
                error = list()
                error.append(codC)
                error.append(des)
                error.append(dataOp)
                error.append(desOp)
                error.append(numDoc)
                error.append(dataDoc)
                error.append(row[4].value)
                error.append(row[5].value)
                sheet.append(error)
                # Salva le modifiche nel file
                workbook.save(file_path)
            else:
                temp = row[4].value
                if isinstance(temp, float):
                    importo = temp
                    temp = 0
                else:
                    temp = temp.strip().replace('.', '').replace(',', '.')
                    x = str(temp)
                    importo = float(x)
                    temp = 0
    elif row[4].value.__class__ == float:
        importo = row[4].value

    # saldo = row[8] # il saldo del conto per ogni operazione
    contr = row[5].value  # la contropartita dell'operazione
    item = ItemConto(codC, des, dataOp, desOp, numDoc, dataDoc, importo, contr)
    # self, codConto, descrConto, dataOp, descrOperazione, numDoc, dataDoc, importo, contropartita):
    return item
