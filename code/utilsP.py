import os

import openpyxl
import xlsxwriter


class ItemConto:
    def __init__(self, codConto, descrConto, dataOp, descrOperazione, numDoc, dataDoc, importo, contropartita):
        self.codConto = codConto
        self.descrConto = descrConto
        self.dataOp = dataOp
        # self.COD = COD
        self.descrOperazione = descrOperazione
        self.numDoc = numDoc
        self.dataDoc = dataDoc
        # self.numFattura = numFattura
        # self.dare = dare
        # self.avere = avere
        self.importo = importo
        # self.saldo = saldo
        self.contropartita = contropartita
        self.costiDir = None
        self.costiIndir = None
        self.attivEconom = None
        self.attivNonEconom = None
        self.CodProg = None

    # Getter
    def getCodiceConto(self):
        return self.codConto

    def getDescrizioneConto(self):
        return self.descrConto

    def getDataOperazione(self):
        return self.dataOp

    def getCodice(self):
        return self.COD

    def getDescrizioneOperazione(self):
        return self.descrOperazione

    def getNumeroDocumento(self):
        return self.numDoc

    def getDataDocumento(self):
        return self.dataDoc
    '''
    def getNumFattura(self):
        return self.numFattura
'''
    def getImporto(self):
        return self.importo

    def getSaldo(self):
        return self.saldo

    def getContropartita(self):
        return self.contropartita

    def getCosti_Diretti(self):
        return self.costiDir

    def getCosti_Indiretti(self):
        return self.costiIndir

    def getAttivita_Economiche(self):
        return self.attivEconom

    def getAttivita_Non_Economiche(self):
        return self.attivNonEconom

    def setCosti_Diretti(self, bool):
        self.costiDir = bool

    def setCosti_Indiretti(self, bool):
        self.costiIndir = bool

    def setAttivita_Economiche(self, bool):
        self.attivEconom = bool

    def setAttivita_Non_Economiche(self, bool):
        self.attivNonEconom = bool

    def getCodiceProgetto(self):
        return self.CodProg

    """
    def getAvere(self):
        return self.avere
    def getDare(self):
        return self.dare
    """


def toList(obj):
    """
    Restituisce una lista dei valori degli attributi di un oggetto.
    """
    attribute_values = []
    for i in obj.__dict__.values():
        attribute_values.append(i)
    return attribute_values


def writeNewFileseparati(listItem):
    outWorkbook = xlsxwriter.Workbook("out2.xlsx")
    outSheet = outWorkbook.add_worksheet(name=listItem[0].getCodiceConto())

    outSheet.write(0, 0, "Codice Conto")
    outSheet.write(0, 1, "Descrizione conto")
    outSheet.write(0, 2, "Data operazione")
    outSheet.write(0, 3, "COD")
    outSheet.write(0, 4, "Descrizione operazione")
    outSheet.write(0, 5, "Numero documento")
    outSheet.write(0, 6, "Data documento")
    outSheet.write(0, 7, "Numero Fattura")
    outSheet.write(0, 8, "Importo")
    outSheet.write(0, 9, "Saldo")
    outSheet.write(0, 10, "Contropartita")
    outSheet.write(0, 11, "Costi Diretti")
    outSheet.write(0, 12, "Costi Indiretti")
    outSheet.write(0, 13, "Attività economiche")
    outSheet.write(0, 14, "Attività non economiche")
    outSheet.write(0, 15, "Codice progetto")

    j = 0
    temp = None
    for i in range(len(listItem)):
        if temp is None:
            temp = listItem[i].getCodiceConto()
            caricamento(outSheet, listItem, i, j)
            j += 1
        if listItem[i].getCodiceConto() == temp:
            caricamento(outSheet, listItem, i, j)
            j += 1
        else:
            j = -1
            outSheet = outWorkbook.add_worksheet(name=listItem[i].getCodiceConto())
            caricamento(outSheet, listItem, i, j)
            j += 1
            temp = listItem[i].getCodiceConto()
    outWorkbook.close()


def caricamento(outSheet, listItem, i, j):
    outSheet.write(j + 1, 0, listItem[i].getCodiceConto())
    outSheet.write(j + 1, 1, listItem[i].getDescrizioneConto())
    outSheet.write(j + 1, 2, listItem[i].getDataOperazione())
    outSheet.write(j + 1, 3, listItem[i].getCodice())
    outSheet.write(j + 1, 4, listItem[i].getDescrizioneOperazione())
    outSheet.write(j + 1, 5, listItem[i].getNumeroDocumento())
    outSheet.write(j + 1, 6, listItem[i].getDataDocumento())
    outSheet.write(j + 1, 7, listItem[i].getNumFattura())
    outSheet.write(j + 1, 8, listItem[i].getImporto())
    outSheet.write(j + 1, 9, listItem[i].getSaldo())
    outSheet.write(j + 1, 10, listItem[i].getContropartita())


def writeNewFile(listItem):
    filename = "out.xlsx"

    # Controlla se il file esiste
    if os.path.isfile(filename):
        # Rimuove il file
        os.remove(filename)
    else:
        print(f"Il file {filename} non esiste")

    outWorkbook = openpyxl.Workbook()
    # outWorkbook = xlsxwriter.Workbook("out.xlsx")

    outSheet = outWorkbook.active

    # Itera su tutte le righe e colonne e imposta il valore di ogni cella su None
    for row in outSheet.iter_rows():
        for cell in row:
            cell.value = None

    outWorkbook.save(filename)

    outWorkbook = openpyxl.load_workbook(filename)

    outSheet = outWorkbook.active

    colonne = ["Codice Conto", "Descrizione conto", "Data operazione", "Descrizione operazione", "Numero documento", "Data documento", "Importo", "Contropartita",
            "Costi Diretti", "Costi Indiretti", "Attività economiche", "Attività non economiche", "Codice progetto"]
    '''
    outSheet.write(0, 0, "Codice Conto")
    outSheet.write(0, 1, "Descrizione conto")
    outSheet.write(0, 2, "Data operazione")
    # outSheet.write(0, 3, "COD")
    outSheet.write(0, 3, "Descrizione operazione")
    outSheet.write(0, 4, "Numero documento")
    outSheet.write(0, 5, "Data documento")
    # outSheet.write(0, 7, "Numero Fattura")
    outSheet.write(0, 6, "Importo")
    # outSheet.write(0, 9, "Saldo")
    outSheet.write(0, 7, "Contropartita")
    outSheet.write(0, 8, "Costi Diretti")
    outSheet.write(0, 9, "Costi Indiretti")
    outSheet.write(0, 10, "Attività economiche")
    outSheet.write(0, 11, "Attività non economiche")
    outSheet.write(0, 12, "Codice progetto")
    '''
    outSheet.append(colonne)
    '''
    for i in range(len(listItem)):
        outSheet.write(i + 1, 0, listItem[i].getCodiceConto())
        outSheet.write(i + 1, 1, listItem[i].getDescrizioneConto())
        outSheet.write(i + 1, 2, listItem[i].getDataOperazione())
        # outSheet.write(i + 1, 3, listItem[i].getCodice())
        outSheet.write(i + 1, 3, listItem[i].getDescrizioneOperazione())
        outSheet.write(i + 1, 4, listItem[i].getNumeroDocumento())
        outSheet.write(i + 1, 5, listItem[i].getDataDocumento())
        # outSheet.write(i + 1, 7, listItem[i].getNumFattura())
        outSheet.write(i + 1, 6, listItem[i].getImporto())
        # outSheet.write(i + 1, 9, listItem[i].getSaldo())
        outSheet.write(i + 1, 7, listItem[i].getContropartita())
    '''
    for item in listItem:
        values = toList(item)
        # print(values)
        outSheet.append(values)

    outWorkbook.save(filename)
    outWorkbook.close()
